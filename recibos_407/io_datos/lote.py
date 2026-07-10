"""
lote.py — Lectura de CSV/XLSX y procesamiento en lote.

Usa csv (stdlib) y openpyxl (liviano) directamente, sin pandas.
Itera fila por fila para no cargar todo en memoria.
"""

from __future__ import annotations

import csv
import os
from dataclasses import dataclass, field
from typing import Callable, Dict, Iterator, List, Optional

from ..core.motor_calculo import ParametrosLiquidacion, liquidar
from ..pdf.generador_pdf import generar_recibo_pdf
from ..config.configuracion import gremio_actual

# Columnas obligatorias (las demás son opcionales para sección A).
_OBLIGATORIAS = {"sueldo_bruto", "porcentaje_art", "porcentaje_sindicato", "es_pyme"}

# Columnas opcionales para pisar el aporte adicional del gremio fila por
# fila. Si no están, se usa el gremio configurado en la app (gremio_actual()).
_COL_APORTE_PCT = "aporte_adicional_pct"
_COL_APORTE_NOMBRE = "aporte_adicional_nombre"

_VERDADEROS = {"true", "1", "si", "sí", "yes", "y", "pyme", "verdadero", "x"}


def _a_bool_pyme(valor) -> bool:
    if isinstance(valor, bool):
        return valor
    if valor is None:
        return False
    return str(valor).strip().lower() in _VERDADEROS


def _nombre_archivo_seguro(texto: str, indice: int) -> str:
    base = "".join(c if c.isalnum() or c in " _-" else "_" for c in str(texto)).strip()
    base = base.replace(" ", "_") or f"empleado_{indice}"
    return f"recibo_{indice:04d}_{base}.pdf"


def _safe_float(valor, defecto: float = 0.0) -> float:
    if valor is None or str(valor).strip() == "":
        return defecto
    return float(valor)


def _safe_str(valor, defecto: str = "") -> str:
    if valor is None:
        return defecto
    return str(valor).strip()


# ── Iteradores de filas ──────────────────────────────────────────────────

def _iterar_csv(ruta: str) -> Iterator[Dict[str, str]]:
    """Itera un CSV fila por fila como diccionarios."""
    with open(ruta, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        # Normalizar nombres de columna
        reader.fieldnames = [c.strip().lower() for c in reader.fieldnames]
        yield from reader


def _iterar_xlsx(ruta: str) -> Iterator[Dict[str, str]]:
    """Itera un XLSX fila por fila como diccionarios (openpyxl, read-only)."""
    from openpyxl import load_workbook

    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    # Primera fila = encabezados
    encabezados = [str(c).strip().lower() if c else f"col_{i}"
                   for i, c in enumerate(next(rows))]

    for fila in rows:
        yield dict(zip(encabezados, fila))

    wb.close()


def _iterar_archivo(ruta: str) -> Iterator[Dict[str, str]]:
    """Selecciona el iterador correcto según la extensión."""
    ext = os.path.splitext(ruta)[1].lower()
    if ext == ".csv":
        yield from _iterar_csv(ruta)
    elif ext in (".xlsx", ".xls"):
        yield from _iterar_xlsx(ruta)
    else:
        raise ValueError(f"Formato no soportado: {ext}. Use .xlsx o .csv")


def _validar_columnas(primera_fila: Dict) -> None:
    """Verifica que las columnas obligatorias estén presentes."""
    presentes = set(primera_fila.keys())
    faltantes = _OBLIGATORIAS - presentes
    if faltantes:
        raise ValueError(
            f"Faltan columnas obligatorias: {', '.join(sorted(faltantes))}")


# ── Conteo rápido ────────────────────────────────────────────────────────

def _contar_filas(ruta: str) -> int:
    """Cuenta filas sin cargar todo en memoria."""
    ext = os.path.splitext(ruta)[1].lower()
    if ext == ".csv":
        with open(ruta, "r", encoding="utf-8-sig") as f:
            return sum(1 for _ in f) - 1  # menos el encabezado
    elif ext in (".xlsx", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(ruta, read_only=True, data_only=True)
        total = wb.active.max_row - 1  # menos el encabezado
        wb.close()
        return max(total, 0)
    return 0


@dataclass
class ResultadoLote:
    """Resumen de un procesamiento por lotes."""
    total: int = 0
    generados: int = 0
    errores: List[str] = field(default_factory=list)
    archivos: List[str] = field(default_factory=list)


def procesar_lote(
    ruta_entrada: str,
    carpeta_salida: str,
    progreso: Optional[Callable[[int, int], None]] = None,
) -> ResultadoLote:
    """
    Genera un PDF por cada fila del archivo de entrada.

    Args:
        ruta_entrada: archivo .xlsx o .csv.
        carpeta_salida: carpeta destino para los PDFs.
        progreso: callback ``fn(hechos, total)`` para reportar avance.

    El aporte adicional del gremio (ej. FAECYS) se toma del gremio
    configurado en la app (una sola lectura para todo el lote, no por
    fila). Si el archivo trae las columnas opcionales
    ``aporte_adicional_pct`` / ``aporte_adicional_nombre``, pisan ese
    valor fila por fila.

    Returns:
        ResultadoLote con resumen.
    """
    os.makedirs(carpeta_salida, exist_ok=True)
    total = _contar_filas(ruta_entrada)
    res = ResultadoLote(total=total)
    gremio = gremio_actual()

    primera = True
    for i, fila in enumerate(_iterar_archivo(ruta_entrada), start=1):
        if primera:
            _validar_columnas(fila)
            primera = False

        try:
            params = ParametrosLiquidacion(
                sueldo_bruto=_safe_float(fila.get("sueldo_bruto")),
                porcentaje_art=_safe_float(fila.get("porcentaje_art")),
                porcentaje_sindicato=_safe_float(fila.get("porcentaje_sindicato")),
                es_pyme=_a_bool_pyme(fila.get("es_pyme")),
                nombre=_safe_str(fila.get("nombre")),
                cuil=_safe_str(fila.get("cuil")),
                legajo=_safe_str(fila.get("legajo")),
                empleador=_safe_str(fila.get("empleador")),
                cuit_empleador=_safe_str(fila.get("cuit_empleador")),
                periodo=_safe_str(fila.get("periodo")),
                categoria=_safe_str(fila.get("categoria")),
                aporte_adicional_pct=_safe_float(
                    fila.get(_COL_APORTE_PCT), gremio.aporte_adicional_pct),
                aporte_adicional_nombre=_safe_str(
                    fila.get(_COL_APORTE_NOMBRE), gremio.aporte_adicional_nombre),
            )
            resultado = liquidar(params)
            nombre_pdf = _nombre_archivo_seguro(params.nombre, i)
            ruta_pdf = os.path.join(carpeta_salida, nombre_pdf)
            generar_recibo_pdf(resultado, ruta_pdf)
            res.archivos.append(ruta_pdf)
            res.generados += 1
        except Exception as e:
            res.errores.append(f"Fila {i}: {e}")
        finally:
            if progreso:
                progreso(i, total)

    return res
